"""
خدمات التصدير - Export Services
توفر دوال لتصدير التقارير بصيغ مختلفة (Excel, CSV, PDF)
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """خدمة التصدير الرئيسية"""
    
    @staticmethod
    def export_to_excel(data, filename, sheet_name="Report"):
        """
        تصدير البيانات إلى Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # الأنماط
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # إضافة العنوان
        if 'title' in data:
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = data['title']
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal="center")
            start_row = 3
        else:
            start_row = 1
        
        # إضافة الرؤوس
        if 'headers' in data:
            for col_num, header in enumerate(data['headers'], 1):
                cell = ws.cell(row=start_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
        
        # إضافة البيانات
        if 'rows' in data:
            for row_num, row_data in enumerate(data['rows'], start_row + 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    
                    # تنسيق الأرقام
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
        
        # ضبط عرض الأعمدة
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # حفظ في buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # إنشاء الاستجابة
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        return response
    
    @staticmethod
    def export_to_csv(data, filename):
        """
        تصدير البيانات إلى CSV
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # إضافة العنوان
        if 'title' in data:
            writer.writerow([data['title']])
            writer.writerow([])  # سطر فارغ
        
        # إضافة الرؤوس
        if 'headers' in data:
            writer.writerow(data['headers'])
        
        # إضافة البيانات
        if 'rows' in data:
            for row in data['rows']:
                writer.writerow(row)
        
        # إنشاء الاستجابة
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        return response
    
    @staticmethod
    def prepare_profitability_data(report_data):
        """
        تحضير بيانات تقرير الربحية للتصدير
        """
        title = _("تقرير الربحية")
        period = f"{report_data['period']['start']} - {report_data['period']['end']}"
        
        data = {
            'title': f"{title} ({period})",
            'headers': [
                _('البند'),
                _('المبلغ'),
                _('النسبة %')
            ],
            'rows': []
        }
        
        # الإيرادات
        total_revenue = report_data['revenue']['total']
        data['rows'].append([_('إجمالي الإيرادات'), total_revenue, '100.00'])
        data['rows'].append([])  # سطر فارغ
        
        # تفصيل الإيرادات حسب طريقة الدفع
        for method in report_data['revenue']['by_method']:
            percentage = (method['total'] / total_revenue * 100) if total_revenue > 0 else 0
            data['rows'].append([
                f"  - {method['payment_method']}",
                method['total'],
                f"{percentage:.2f}"
            ])
        
        data['rows'].append([])  # سطر فارغ
        
        # المصروفات
        total_expenses = report_data['expenses']['total']
        expense_percentage = (total_expenses / total_revenue * 100) if total_revenue > 0 else 0
        data['rows'].append([_('إجمالي المصروفات'), total_expenses, f"{expense_percentage:.2f}"])
        data['rows'].append([])  # سطر فارغ
        
        # تفصيل المصروفات حسب الفئة
        for category in report_data['expenses']['by_category']:
            percentage = (category['total'] / total_revenue * 100) if total_revenue > 0 else 0
            data['rows'].append([
                f"  - {category['category']}",
                category['total'],
                f"{percentage:.2f}"
            ])
        
        data['rows'].append([])  # سطر فارغ
        
        # صافي الربح
        net_profit = report_data['profitability']['net_profit']
        profit_margin = report_data['profitability']['profit_margin']
        data['rows'].append([_('صافي الربح'), net_profit, f"{profit_margin:.2f}"])
        
        return data
    
    @staticmethod
    def prepare_cash_flow_data(report_data):
        """
        تحضير بيانات تقرير التدفق النقدي للتصدير
        """
        title = _("تقرير التدفق النقدي")
        period = f"{report_data['period']['start']} - {report_data['period']['end']}"
        
        data = {
            'title': f"{title} ({period})",
            'headers': [
                _('الشهر'),
                _('التدفقات الداخلة'),
                _('التدفقات الخارجة'),
                _('صافي التدفق')
            ],
            'rows': []
        }
        
        for item in report_data['monthly_cash_flow']:
            month_name = item['month'].strftime('%Y-%m')
            data['rows'].append([
                month_name,
                item['inflow'],
                item['outflow'],
                item['net_flow']
            ])
        
        # الإجماليات
        data['rows'].append([])  # سطر فارغ
        data['rows'].append([
            _('الإجمالي'),
            report_data['totals']['inflow'],
            report_data['totals']['outflow'],
            report_data['totals']['net_flow']
        ])
        
        return data
    
    @staticmethod
    def prepare_occupancy_data(report_data):
        """
        تحضير بيانات تقرير معدل الإشغال للتصدير
        """
        title = _("تقرير معدل الإشغال")
        date = report_data['date']
        
        data = {
            'title': f"{title} ({date})",
            'headers': [
                _('البند'),
                _('إجمالي الوحدات'),
                _('المشغول'),
                _('الشاغر'),
                _('معدل الإشغال %')
            ],
            'rows': []
        }
        
        # الملخص العام
        summary = report_data['summary']
        data['rows'].append([
            _('الإجمالي'),
            summary['total_units'],
            summary['occupied_units'],
            summary['vacant_units'],
            f"{summary['occupancy_rate']:.2f}"
        ])
        
        data['rows'].append([])  # سطر فارغ
        
        # حسب نوع الوحدة
        data['rows'].append([_('حسب نوع الوحدة'), '', '', '', ''])
        for item in report_data['by_type']:
            data['rows'].append([
                f"  - {item['unit_type']}",
                item['total'],
                item['occupied'],
                item['vacant'],
                f"{item['occupancy_rate']:.2f}"
            ])
        
        data['rows'].append([])  # سطر فارغ
        
        # حسب المبنى
        data['rows'].append([_('حسب المبنى'), '', '', '', ''])
        for item in report_data['by_building']:
            data['rows'].append([
                f"  - {item['building__name']}",
                item['total'],
                item['occupied'],
                item['vacant'],
                f"{item['occupancy_rate']:.2f}"
            ])
        
        return data
    
    @staticmethod
    def prepare_overdue_tenants_data(report_data):
        """
        تحضير بيانات تقرير المستأجرين المتأخرين للتصدير
        """
        title = _("تقرير المستأجرين المتأخرين")
        date = report_data['as_of_date']
        
        data = {
            'title': f"{title} ({date})",
            'headers': [
                _('المستأجر'),
                _('الهاتف'),
                _('الوحدة'),
                _('رقم العقد'),
                _('الإيجار الشهري'),
                _('عدد الشهور المتأخرة'),
                _('المبلغ المتأخر'),
                _('أقصى تأخير (أيام)')
            ],
            'rows': []
        }
        
        for tenant in report_data['tenants']:
            data['rows'].append([
                tenant['tenant_name'],
                tenant['tenant_phone'],
                tenant['unit'],
                tenant['lease_number'],
                tenant['monthly_rent'],
                tenant['overdue_months_count'],
                tenant['total_overdue_amount'],
                tenant['max_days_overdue']
            ])
        
        # الإجماليات
        data['rows'].append([])  # سطر فارغ
        data['rows'].append([
            _('الإجمالي'),
            '',
            '',
            '',
            '',
            report_data['summary']['total_overdue_months'],
            report_data['summary']['total_overdue_amount'],
            ''
        ])
        
        return data
    
    @staticmethod
    def prepare_lease_expiry_data(report_data):
        """
        تحضير بيانات تقرير انتهاء العقود للتصدير
        """
        title = _("تقرير انتهاء العقود")
        period = f"{report_data['period']['start']} - {report_data['period']['end']}"
        
        data = {
            'title': f"{title} ({period})",
            'headers': [
                _('رقم العقد'),
                _('المستأجر'),
                _('الهاتف'),
                _('الوحدة'),
                _('تاريخ البداية'),
                _('تاريخ الانتهاء'),
                _('الأيام المتبقية'),
                _('الإيجار الشهري'),
                _('إجمالي الإيجار')
            ],
            'rows': []
        }
        
        for lease in report_data['leases']:
            data['rows'].append([
                lease['lease_number'],
                lease['tenant_name'],
                lease['tenant_phone'],
                lease['unit'],
                lease['start_date'],
                lease['end_date'],
                lease['days_until_expiry'],
                lease['monthly_rent'],
                lease['total_rent']
            ])
        
        return data


class ReportExporter:
    """مُصدّر التقارير"""
    
    @staticmethod
    def export_report(report_type, report_data, format='excel', filename=None):
        """
        تصدير تقرير بصيغة محددة
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{report_type}_{timestamp}"
        
        # تحضير البيانات حسب نوع التقرير
        if report_type == 'profitability':
            data = ExportService.prepare_profitability_data(report_data)
        elif report_type == 'cash_flow':
            data = ExportService.prepare_cash_flow_data(report_data)
        elif report_type == 'occupancy':
            data = ExportService.prepare_occupancy_data(report_data)
        elif report_type == 'overdue_tenants':
            data = ExportService.prepare_overdue_tenants_data(report_data)
        elif report_type == 'lease_expiry':
            data = ExportService.prepare_lease_expiry_data(report_data)
        else:
            raise ValueError(f"Unsupported report type: {report_type}")
        
        # التصدير بالصيغة المطلوبة
        if format == 'excel':
            return ExportService.export_to_excel(data, filename)
        elif format == 'csv':
            return ExportService.export_to_csv(data, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")
